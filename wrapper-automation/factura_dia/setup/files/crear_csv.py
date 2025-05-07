import pandas as pd
from openpyxl import load_workbook

# Ruta del archivo Excel
file_excel = 'facturas-dia.xlsx'
file_csv = 'facturas-dia.csv'

# Cargar libro y hoja activa
wb = load_workbook(file_excel)
ws = wb.active

# Detectar celdas combinadas
merged_cells = ws.merged_cells.ranges

# Función para detectar si una celda está combinada
def is_merged(cell, merged_ranges):
    return any(cell.coordinate in merged_range for merged_range in merged_ranges)

# Fila de encabezado (fila 8), pero no se incluye en los datos
header_row_num = 8
header_row = ws[header_row_num]

# Generar nombres de columnas usando fila 8
columns = []
for cell in header_row:
    col_name = cell.value
    if col_name is None or is_merged(cell, merged_cells):
        col_name = f"Col{cell.column:02d}"
    columns.append(col_name)

# Leer los datos desde la fila 9
# skiprows=8 → salta filas 1 a 8
df = pd.read_excel(file_excel, skiprows=8, header=None)

# Asignar los nombres de columna generados previamente
df.columns = columns[:len(df.columns)]  # Evita error si hay menos columnas

# Eliminar filas completamente vacías (opcional)
df.dropna(how='all', inplace=True)

# Guardar a CSV con codificación UTF-8-SIG (compatibilidad con Excel)
df.to_csv(file_csv, index=False, sep=';', encoding='utf-8-sig')

print("✅ Archivo CSV generado correctamente desde la fila 9 sin duplicar encabezados.")