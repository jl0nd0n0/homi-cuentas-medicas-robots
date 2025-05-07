from dotenv import load_dotenv
import os
import sys
# Get the absolute path of the parent folder
parent_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Add the parent folder to sys.path
sys.path.append(parent_folder)
import pyautogui
from core import robotClick
import time
from pywinauto import Application
from datetime import datetime
import uiautomation as auto
import mysql.connector
import subprocess

import pandas as pd
from openpyxl import load_workbook

class HomiRobotFacturaDia:
    """
    Class to handle the robot for generating daily invoices
    """

    def __init__(self, dia, mes, año, script_dir):
        #path = r"C:\tools\robot\homi\factura_dia"
        os.chdir(script_dir)
        print(script_dir)
        file = script_dir + rf"\facturas-dia.xlsx"
        #print(file)
        #sys.exit()

        # Cargar variables de entorno desde el archivo .env
        # Obtener la ruta absoluta del archivo .env en el directorio superior
        env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env')
        # print(env_path)
        load_dotenv(env_path)

        # Configuración de la conexión a la base de datos usando variables de entorno
        db_config = {
            "host": os.getenv("DB_HOST"),
            "user": os.getenv("DB_USER"),
            "password": os.getenv("DB_PASSWORD"),
            "database": os.getenv("DB_NAME")
        }

        def actualizar_generado(factura):
            """
            Actualiza el campo 'generado' a 1 para la factura especificada.
            
            :param factura: El número de factura a actualizar.
            """
            try:
                # Conectar a la base de datos
                connection = mysql.connector.connect(**db_config)
                cursor = connection.cursor()

                # Consulta SQL para actualizar el campo 'generado'
                query = """
                    UPDATE soporte_generar
                    SET generado = 1
                    WHERE numero_factura = %s
                    and soporte = 'factura-excel'
                """
                # Ejecutar la consulta con el parámetro de la factura
                cursor.execute(query, (factura,))
                connection.commit()

                # Verificar si se realizó la actualización
                if cursor.rowcount > 0:
                    print(f"El campo 'generado' ha sido actualizado a 1 para la factura {factura}.")
                else:
                    print(f"No se encontró ninguna factura con el número {factura}.")

            except mysql.connector.Error as err:
                print(f"Error al conectar o actualizar en la base de datos: {err}")
            finally:
                # Cerrar la conexión a la base de datos
                if 'connection' in locals() and connection.is_connected():
                    cursor.close()
                    connection.close()
                    print("Conexión a la base de datos cerrada.")

        print(str(dia) + " " + str(mes) + " " + str(año))
        #sys.exit()

        dia = int(dia)
        mes = int(mes)
        año = int(año)

        dias_mes_array = [31, "b", 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        # recuperar los dias por mes el array inicia en 0, para el primero enero, hay que restar 1
        dias_mes = dias_mes_array[mes-1]
        #print(str(dias_mes));

        # si el año es bisiesto, febrero tiene 29 dias de lo contrario 28
        if (mes == 2):
            residuo = año % 4
            if (residuo == 0):
                dias_mes = 29
            else:
                dias_mes = 28

        # si es el ultimo dia del mes, el dia siguiente es el primero del nuevo mes 
        if (dia == dias_mes):
            dia_siguiente = 1
            mes_siguiente = mes + 1
        else:
            dia_siguiente = dia + 1
            mes_siguiente = mes
        #print("dia siguiente " + str(dia_siguiente))


        # si es el ultimo dia (31) del ultimo mes (12) el año es el siguiente 
        if (dia == 31 and mes == 12):
            año_siguiente = año + 1
        else:
            año_siguiente = año

        # formato dia, dia siguiente, mes y año para que el robot digite
        if (len(str(dia)) == 1):
            dia = "0" + str(dia)
        else:
            dia = str(dia)

        if (len(str(dia_siguiente)) == 1):
            dia_siguiente = "0" + str(dia_siguiente)
        else:
            dia_siguiente = str(dia_siguiente)
            
        if (len(str(mes)) == 1):
            mes = "0" + str(mes)
        else:
            mes = str(mes)
            
        if (len(str(mes_siguiente)) == 1):
            mes_siguiente = "0" + str(mes_siguiente)
        else:
            mes_siguiente = str(mes_siguiente)
            
        año = str(año)
        año_siguiente = str(año_siguiente)

        print(dia + "," + dia_siguiente + "," + mes + "," + año)
        #sys.exit()

        # Find a window by its title
        window = auto.WindowControl(searchDepth=1, AutomationId="FormMdi")
        if  not window.Exists():
            print("La ventana no existe !!!")
        window.SetActive()

        # Send Ctrl + F4
        pyautogui.hotkey('ctrl', 'f4')

        # (1)
        robotClick(83, 253, 1, "Click Opcion vie RCM")
        # (2)
        robotClick(317, 162, 1, "Click Facturacion en Salud")
        # (3)
        robotClick(693, 481, 3, "Click Lista de Facturas")
        # (4)
        robotClick(237, 180, 1, "Click Input Fecha Inicial")
        pyautogui.typewrite(dia)
        pyautogui.typewrite(mes)
        pyautogui.typewrite(año)

        # (5)
        robotClick(233, 235, 0, "Click Input Fecha Final")
        #pyautogui.typewrite(str(dia + 2))
        pyautogui.typewrite(dia_siguiente)
        pyautogui.typewrite(mes_siguiente)
        pyautogui.typewrite(año_siguiente)
        #sys.exit()

        # (6)
        messageButton = "Generar Reporte"
        print("click boton " + messageButton)
        button = auto.Control(Name = "Generar Reporte")
        if not button.Exists():
            print("No se ha encontrado el boton " + messageButton)
            sys.exit()
        button.Click()
        time.sleep(4)

        # (7)
        '''
        messageButton = "Export Document as ..."
        print("click boton " + messageButton)
        button = window.Control(Name = "Export Document...")
        if not button.Exists():
            print("No se ha encontrado el boton " + messageButton)
            sys.exit()
        button.Click()
        time.sleep(1)
        buttonChild = button.Control(Name="Elemento")
        #buttonExportarComo = button.Control(Name = "Elemento")
        '''
        messageButton = "Export Document as ..."
        print("click boton " + messageButton)
        button = window.Control(Name = "Export Document...")
        if not button.Exists():
            print("No se ha encontrado el boton " + messageButton)
            sys.exit()
        # Get the bounding rectangle of the button
        rect = button.BoundingRectangle
        # Calculate center x and y
        center_x = (rect.left + rect.right) // 2
        center_y = (rect.top + rect.bottom) // 2
        #click in child button export as types ....
        robotClick(center_x + 14, center_y, 0, "Click Input Fecha Final")
        time.sleep(1)

        # (8) 
        button = auto.Control(Name="XLSX File")
        if not button.Exists():
            print("No se ha encontrado el boton XLSX File")
            sys.exit()
        button.Click()
        time.sleep(1)

        # (9)
        panel = auto.Control(Name="The XtraLayoutControl")
        if not panel.Exists():
            print("No se ha encontrado el panel de opciones de exportacion XLSX")
            sys.exit()
        button = panel.Control(Name="Aceptar")
        if not button.Exists():
            print("No se ha encontrado el boton aceptar panel opciones de exportacion XLSX")
            sys.exit()
        button.Click()
        time.sleep(1)

        # (10)
        # Crea el directorio si no existe
        pyautogui.typewrite(file)
        auto.SendKeys("{Enter}")
        time.sleep(1)

        # (11)
        print("presionando alt + g");
        # presional alt + g
        auto.SendKeys('%g')
        #esperar un segundo 
        time.sleep(1)
        # presional alt + s
        #si existe el archivo volver a guardarlo
        dialogo = window.Control(Name="Guardar como")
        if dialogo.Exists():   
            button = dialogo.Control(AutomationId="CommandButton_6")
            if not button.Exists():
                print("No se ha encontrado el button Guardar Como")
                sys.exit()
            button.Click()
        time.sleep(1)

        # (12)
        print("presionando alt + n");
        auto.SendKeys('%n')
        time.sleep(1)

        # (13)
        print("click boton Exit (rojo x)");
        button = window.ButtonControl(Name = "Exit")
        if not button.Exists():
            print("No se ha encontrado el boton Exit")
            sys.exit()
        button.Click()
        subprocess.run(["python", "crear_csv.py"])

        # load data to database
        # Path to your .bat file
        #bat_file_path = "robot.bat"
        # Run the .bat file
        #subprocess.run(bat_file_path, shell=True)

        # Name of your .bat file
        bat_file = script_dir + r"\robot.bat"
        #print(bat_file)
        #sys.exit()

        # Run the batch file
        result = subprocess.run(
            [bat_file],          # Command to run
            shell=True,          # Required for .bat files
            check=True           # Raise exception if it fails
        )
        os.chdir("..")
        print(f"Batch file exited with code {result.returncode}")

    def crearCSV(self, file, script_dir):
        os.chdir(script_dir)
        print("Current Working Directory:", os.getcwd())
        # Ruta del archivo Excel

        script_dir = os.path.dirname(os.path.abspath(__file__))

        file_excel = 'facturas-dia.xlsx'
        file_csv = 'facturas-dia.csv'

        # Cargar libro y hoja activa
        wb = load_workbook(os.path.join(script_dir, file_excel))
        ws = wb.active

        # Detectar celdas combinadas
        merged_cells = ws.merged_cells.ranges

        # Función para detectar si una celda está combinada
        def is_merged(cell, merged_ranges):
            return any(cell.coordinate in merged_range for merged_range in merged_ranges)

        # Fila de encabezado (fila 8), pero no se incluye en los datos
        header_row_num = 8
        header_row = ws[header_row_num]

        # Generar nombres de columnas usando fila 8
        columns = []
        for cell in header_row:
            col_name = cell.value
            if col_name is None or is_merged(cell, merged_cells):
                col_name = f"Col{cell.column:02d}"
            columns.append(col_name)

        # Leer los datos desde la fila 9
        # skiprows=8 → salta filas 1 a 8
        df = pd.read_excel(file_excel, skiprows=8, header=None)

        # Asignar los nombres de columna generados previamente
        df.columns = columns[:len(df.columns)]  # Evita error si hay menos columnas

        # Eliminar filas completamente vacías (opcional)
        df.dropna(how='all', inplace=True)

        # Guardar a CSV con codificación UTF-8-SIG (compatibilidad con Excel)
        df.to_csv(os.path.join(script_dir, file_csv), index=False, sep=';', encoding='utf-8-sig')
        os.chdir("..")
        print("✅ Archivo CSV generado correctamente desde la fila 9 sin duplicar encabezados.")